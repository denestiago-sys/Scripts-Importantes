import re
import openpyxl
import pdfplumber
import io
from io import BytesIO
from pathlib import Path

# --- PADRÕES DE BUSCA (REGEX) ---
META_RE  = re.compile(r"^META ESPEC[ÍI]FICA\s+(\d+)", re.IGNORECASE)
ITEM_RE  = re.compile(r"^Item\s*(\d+)\s*(Planejado|Aprovado|Cancelado)?", re.IGNORECASE)
ART_PATTERN = re.compile(r"^Art\.?\s*(6|7|8)\s*º?\s*(?:\((\d+)\))?\s*:\s*(.*)", re.IGNORECASE)

# Padrão para isolar apenas o valor em dinheiro (R$ 0.000,00)
VALOR_MONETARIO_RE = re.compile(r"(R\$\s?[\d\.,]+)")

# Linhas que encerram a captura do campo atual (não pertencem a nenhum campo)
STOP_CAPTURE_RE = re.compile(
    r"^(Valor Origin[aá]rio Planejado"
    r"|Valor Suplementar Planejado"
    r"|Valor Rendimento Planejado"
    r"|C[oó]d\.?\s*Senasp)",
    re.IGNORECASE
)

RUÍDO_RODAPÉ = [
    re.compile(r"https?://apps\.mj\.gov\.br", re.IGNORECASE),
    re.compile(r"Planos de Aplicação",         re.IGNORECASE),
    re.compile(r"\d{2}/\d{2}/\d{4},? \d{2}:\d{2}"),
    re.compile(r"\d+\s*/\s*\d+"),
    re.compile(r"CÓDIGO DE VERIFICAÇÃO",        re.IGNORECASE),
    re.compile(r"Reduzir em \d+%",              re.IGNORECASE),
    re.compile(r"Descrição do Indicador",       re.IGNORECASE),
    # Linhas de valores intermediários que não devem ser capturadas
    re.compile(r"^Valor Origin[aá]rio Planejado:", re.IGNORECASE),
    re.compile(r"^Valor Suplementar Planejado:",   re.IGNORECASE),
    re.compile(r"^Valor Rendimento Planejado:",    re.IGNORECASE),
    re.compile(r"^C[oó]d\.?\s*Senasp:",           re.IGNORECASE),
]

CAPTURE_PATTERNS = [
    ("bem",         re.compile(r"^(?:Bem|Material)/Servi[cç]o:\s*(.*)", re.IGNORECASE)),
    ("descricao",   re.compile(r"^Descri[cç][aã]o:\s*(.*)",            re.IGNORECASE)),
    ("destinacao",  re.compile(r"^Destina[cç][aã]o:\s*(.*)",           re.IGNORECASE)),
    ("unidade",     re.compile(r"^Unidade de Medida:\s*(.*)",           re.IGNORECASE)),
    ("quantidade",  re.compile(r"^Qtd\.?\s*Planejada:\s*(.*)",         re.IGNORECASE)),
    ("natureza",    re.compile(r"^Natureza\s*\(ND\):\s*(.*)",          re.IGNORECASE)),
    ("instituicao", re.compile(r"^Institui[cç][aã]o:\s*(.*)",          re.IGNORECASE)),
    ("valor_total", re.compile(r"^Valor Total:\s*(.*)",                 re.IGNORECASE)),
]


# ---------------------------------------------------------------------------
# Funções auxiliares
# ---------------------------------------------------------------------------

def limpar_valor_financeiro(texto):
    """Extrai apenas o último padrão de R$ encontrado na linha."""
    matches = VALOR_MONETARIO_RE.findall(texto)
    if matches:
        return matches[-1]
    return texto


def eh_ruido(texto):
    return any(p.search(texto) for p in RUÍDO_RODAPÉ)


def normalize(text):
    if "https://" in text:
        text = text.split("https://")[0]
    return re.sub(r"\s+", " ", text or "").strip()


# ---------------------------------------------------------------------------
# Extração de linhas do PDF
# ---------------------------------------------------------------------------

def extract_lines_from_pdf(file_obj):
    file_obj.seek(0)
    lines = []
    with pdfplumber.open(file_obj) as pdf:
        # Aviso se PDF parece ser imagem escaneada
        total_chars = sum(len(page.extract_text() or "") for page in pdf.pages)
        if total_chars < 100:
            raise ValueError(
                "O PDF parece ser uma imagem escaneada. "
                "Use um PDF com texto selecionável."
            )
        # Re-abre para iterar (pdfplumber fecha o contexto ao sair do with)
        file_obj.seek(0)

    file_obj.seek(0)
    with pdfplumber.open(file_obj) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            text = text.replace("\x0c", "\n")
            for raw in text.splitlines():
                if raw.strip() and not eh_ruido(raw):
                    lines.append(raw.strip())
    return lines


# ---------------------------------------------------------------------------
# Parser de itens
# ---------------------------------------------------------------------------

def parse_items(lines):
    items        = []
    current_meta = "N/A"
    current_item_obj = None
    seen_items   = set()

    for line in lines:
        m_meta = META_RE.match(line)
        if m_meta:
            current_meta = m_meta.group(1)
            continue

        m_item = ITEM_RE.match(line)
        if m_item:
            item_num   = m_item.group(1)
            unique_key = f"M{current_meta}I{item_num}"
            if unique_key not in seen_items:
                current_item_obj = {
                    "meta_num": current_meta,
                    "item_num": item_num,
                    "status":   m_item.group(2) or "Planejado",
                    "lines":    []
                }
                items.append(current_item_obj)
                seen_items.add(unique_key)
            continue

        if current_item_obj:
            current_item_obj["lines"].append(line)

    return items


# ---------------------------------------------------------------------------
# Extração de campos de cada item
# ---------------------------------------------------------------------------

def extract_fields(item_lines):
    fields = {key: [] for key, _ in CAPTURE_PATTERNS}
    fields["art_texto"] = ""
    current_f = None

    for line in item_lines:
        # Artigo
        if ART_PATTERN.match(line):
            fields["art_texto"] = line
            current_f = None
            continue

        # Linhas que encerram o campo atual (ex.: Valor Originário Planejado)
        if STOP_CAPTURE_RE.match(line):
            current_f = None
            continue

        # Tenta casar com algum campo capturável
        matched = False
        for key, pat in CAPTURE_PATTERNS:
            m = pat.match(line)
            if m:
                fields[key].append(m.group(1))
                current_f = key
                matched = True
                break

        # Continuação multilinha do campo anterior
        if not matched and current_f:
            if not eh_ruido(line):
                fields[current_f].append(line)

    return {
        k: normalize(" ".join(v)) if isinstance(v, list) else v
        for k, v in fields.items()
    }


# ---------------------------------------------------------------------------
# Montagem das linhas para o Excel
# ---------------------------------------------------------------------------

def build_rows(parsed_items):
    rows = []
    for item in parsed_items:
        f = extract_fields(item["lines"])

        valor_limpo   = limpar_valor_financeiro(f["valor_total"])
        unidade_limpa = f["unidade"].split(" http")[0].split(" Reduzir")[0].strip()
        natureza_limpa = f["natureza"].split(" http")[0].strip()

        rows.append({
            "Número da Meta Específica":                    item["meta_num"],
            "Número do Item":                               item["item_num"],
            "Ação conforme Art. 7º da portaria nº 685":    f["art_texto"],
            "Material/Serviço":                             f["bem"],
            "Descrição":                                    f["descricao"],
            "Destinação":                                   f["destinacao"],
            "Instituição":                                  f["instituicao"],
            "Natureza da Despesa":                          natureza_limpa,
            "Quantidade Planejada":                         f["quantidade"],
            "Unidade de Medida":                            unidade_limpa,
            "Valor Planejado Total":                        valor_limpo,
            "Status do Item":                               item["status"],
        })
    return rows


# ---------------------------------------------------------------------------
# Geração do Excel a partir do template
# ---------------------------------------------------------------------------

def generate_excel_bytes(template_path, rows):
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Localiza a linha de cabeçalho
    header_row = 1
    for r in range(1, 10):
        val = str(ws.cell(r, 1).value or "")
        if "Meta" in val or "Número" in val:
            header_row = r
            break

    col_map = {
        str(ws.cell(header_row, c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(header_row, c).value
    }

    for idx, data in enumerate(rows, start=header_row + 1):
        for header_name, col_idx in col_map.items():
            if header_name in data:
                ws.cell(row=idx, column=col_idx, value=data[header_name])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
