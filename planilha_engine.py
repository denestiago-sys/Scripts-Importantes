#!/usr/bin/env python3
import argparse
import copy
import re
from io import BytesIO
from pathlib import Path

import pdfplumber
import openpyxl
from openpyxl.cell.cell import MergedCell

META_RE = re.compile(r"^META ESPEC[ÍI]FICA\s+(\d+)", re.IGNORECASE)
ITEM_RE = re.compile(
    r"^Item\s*(\d+)\s*(Planejado|Aprovado|Cancelado)?", re.IGNORECASE
)
ACTION_HEADER_KEY = "acao_art"
ACTION_HEADER_NUM_KEY = "acao_art_num"
REQUIRED_TEMPLATE_NAME = "Planilha Base(atualizada).xlsx"
ACTION_HEADER_PATTERN = re.compile(
    r"^Ação conforme Art\.\s*\d+º\s+da portaria nº 685$",
    re.IGNORECASE,
)
# Texto de exemplo que o portal insere no campo "Indicador Geral de Resultado"
# (começa com "EX:" e termina com ")"). Deve ser removido do resultado extraído.
EXAMPLE_BLOCK_RE = re.compile(r"EX:\s*\(.*?\)", re.DOTALL | re.IGNORECASE)

PLAN_SIGNATURE_RE = re.compile(
    r"\b([A-Z]{2})\s*-\s*([A-Z0-9]+)\s*-\s*(20\d{2})\b"
)
ART_PATTERN = re.compile(
    r"^Art\.?\s*(6|7|8)\s*º?\s*(?:\((\d+)\))?\s*:\s*(.*)",
    re.IGNORECASE,
)
ACTION_PATTERN = re.compile(r"^A[cç][aã]o:\s*(.*)", re.IGNORECASE)

CAPTURE_PATTERNS = [
    ("bem", re.compile(r"^(?:Bem|Material)/Servi[cç]o:\s*(.*)", re.IGNORECASE)),
    ("descricao", re.compile(r"^Descri[cç][aã]o:\s*(.*)", re.IGNORECASE)),
    ("destinacao", re.compile(r"^Destina[cç][aã]o:\s*(.*)", re.IGNORECASE)),
    ("unidade", re.compile(r"^Unidade de Medida:\s*(.*)", re.IGNORECASE)),
    ("quantidade", re.compile(r"^Qtd\.?\s*Planejada:\s*(.*)", re.IGNORECASE)),
    ("quantidade", re.compile(r"^Quantidade Planejada:\s*(.*)", re.IGNORECASE)),
    ("natureza", re.compile(r"^Natureza\s*\(ND\):\s*(.*)", re.IGNORECASE)),
    ("instituicao", re.compile(r"^Institui[cç][aã]o:\s*(.*)", re.IGNORECASE)),
    ("valor_total", re.compile(r"^Valor Total:\s*(.*)", re.IGNORECASE)),
]

STOP_PATTERNS = [
    re.compile(r"^C[oó]d\.?\s*Senasp:", re.IGNORECASE),
    re.compile(r"^Valor Origin[aá]rio Planejado:", re.IGNORECASE),
    re.compile(r"^Valor Suplementar Planejado:", re.IGNORECASE),
    re.compile(r"^Valor Rendimento Planejado:", re.IGNORECASE),
]

OUTPUT_HEADERS = [
    "Número da Meta Específica",
    "Número do Item",
    "Ação conforme Art. 7º da portaria nº 685",
    "Material/Serviço",
    "Descrição",
    "Destinação",
    "Instituição",
    "Natureza da Despesa",
    "Quantidade Planejada",
    "Unidade de Medida",
    "Valor Planejado Total",
    "Status do Item",
]

ANALYSIS_TEMPLATE_TITLE = "ANÁLISE DOS ELEMENTOS DO PLANO DE APLICAÇÃO"
ANALYSIS_BLOCK_START_ROW = 14
ANALYSIS_BLOCK_HEIGHT = 11
ANALYSIS_BLOCK_START_COL = 1  # A
ANALYSIS_BLOCK_END_COL = 12   # L


def normalize(text: str) -> str:
    text = re.sub(r"\s+", " ", text or "").strip()
    return text


def blank_if_dash_only(value: str) -> str:
    text = normalize(value)
    if re.fullmatch(r"[-–—]+", text):
        return ""
    return text


def strip_currency(value: str) -> str:
    value = (value or "").replace("R$", "").strip()
    value = value.replace(".", "")
    value = re.sub(r"\s+", "", value)
    return value


def format_currency(value: str) -> str:
    value = strip_currency(value)
    if not value:
        return ""
    if "," in value:
        integer_part, decimal_part = value.split(",", 1)
    else:
        integer_part, decimal_part = value, "00"
    integer_part = re.sub(r"[^0-9]", "", integer_part)
    decimal_part = re.sub(r"[^0-9]", "", decimal_part)[:2].ljust(2, "0")
    integer_part = integer_part.lstrip("0") or "0"
    grouped = ""
    while integer_part:
        grouped = integer_part[-3:] + (f".{grouped}" if grouped else "")
        integer_part = integer_part[:-3]
    return f"R$ {grouped},{decimal_part}"


def parse_int(value: str):
    digits = re.sub(r"[^0-9]", "", value or "")
    return int(digits) if digits else ""


def normalize_pdf_text(text: str) -> str:
    text = text.replace("\x0c", "\n")
    text = re.sub(
        r"(META ESPEC[ÍI]FICA\s+\d+)", r"\n\1\n", text, flags=re.IGNORECASE
    )
    # Requer a palavra de status (Planejado/Aprovado/Cancelado) para reconhecer
    # um cabeçalho de item. Sem o status, textos como "REMANEJAMENTO DE SALDO
    # DO ITEM 30" que aparecem em descrições seriam incorretamente tratados como
    # início de um novo item.
    text = re.sub(
        r"(Item\s*\d+\s+(?:Planejado|Aprovado|Cancelado))",
        r"\n\1\n",
        text,
        flags=re.IGNORECASE,
    )
    return text


def clean_lines(text: str):
    lines = []
    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue
        if re.match(r"^\d{2}/\d{2}/\d{4},", line):
            continue
        if "Planos de Aplicação" in line and re.search(r"\d{2}/\d{2}/\d{4}", line):
            continue
        if line.startswith("https://apps.mj.gov.br/"):
            continue
        lines.append(line)
    return lines


def extract_lines_from_pdf(pdf_path: Path):
    lines = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            text = normalize_pdf_text(page.extract_text() or "")
            lines.extend(clean_lines(text))
    return lines


def extract_lines_from_pdf_file(file_obj):
    file_obj.seek(0)
    lines = []
    with pdfplumber.open(file_obj) as pdf:
        for page in pdf.pages:
            text = normalize_pdf_text(page.extract_text() or "")
            lines.extend(clean_lines(text))
    return lines


def extract_plan_signature(lines):
    max_lines = min(len(lines), 120)
    for idx in range(max_lines):
        line = (lines[idx] or "").strip()
        if not line:
            continue
        match = PLAN_SIGNATURE_RE.search(line.upper())
        if match:
            return {
                "sigla": match.group(2).upper(),
                "ano": int(match.group(3)),
                "raw_line": line,
            }
    return {"sigla": None, "ano": None, "raw_line": None}


def resolve_art_by_plan_rule(sigla, ano):
    if not sigla or not ano:
        return None
    sigla = str(sigla).upper()
    if sigla in {"ECV", "FISPDS", "RMVI"} and 2019 <= ano <= 2025:
        return "6"
    if sigla == "EVM" and 2023 <= ano <= 2025:
        return "7"
    if sigla in {"VPSP", "MQVPSP"} and 2019 <= ano <= 2025:
        return "8"
    return None


def is_analysis_template_sheet(ws) -> bool:
    title = normalize(str(ws["A2"].value or "")).upper()
    return ANALYSIS_TEMPLATE_TITLE in title


def is_analysis_template_file(template_path: Path) -> bool:
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    return is_analysis_template_sheet(ws)


def _header_key(header: str) -> str:
    return ACTION_HEADER_KEY if ACTION_HEADER_PATTERN.match(header) else header


def get_header_info_from_ws(ws, header_row: int):
    headers = []
    header_map = {}
    for cell in ws[header_row]:
        if cell.value:
            header = str(cell.value).strip()
            headers.append(header)
            key = _header_key(header)
            if key not in header_map:
                header_map[key] = cell.column
    return headers, header_map


# Campos financeiros que só aparecem APÓS a descrição do item no PDF.
# Usados para distinguir um cabeçalho de item real de uma referência
# a um item dentro de uma descrição (ex.: "REMANEJAMENTO DE SALDO DO ITEM 30").
_FINANCIAL_FIELD_RE = re.compile(
    r"^(Destina[cç][aã]o|Institui[cç][aã]o|Natureza|Valor\s+Total|Qtd\.?|Quantidade)",
    re.IGNORECASE,
)


def _item_has_financial_data(lines) -> bool:
    """Retorna True se alguma linha já coletada é um campo financeiro do item."""
    return any(_FINANCIAL_FIELD_RE.match(l) for l in lines)


def parse_items(lines):
    items = []
    current_meta = None
    current_item = None
    current_status = None
    current_lines = []

    def flush():
        nonlocal current_item, current_lines, current_status
        if current_meta is None or current_item is None:
            return
        items.append({
            "meta": current_meta,
            "item": current_item,
            "status": current_status or "",
            "lines": current_lines[:],
        })
        current_item = None
        current_status = None
        current_lines = []

    for line in lines:
        meta_match = META_RE.match(line)
        if meta_match:
            flush()
            current_meta = int(meta_match.group(1))
            continue
        item_match = ITEM_RE.match(line)
        if item_match:
            new_item   = int(item_match.group(1))
            new_status = (item_match.group(2) or "").capitalize()

            # Guarda de segurança: "Item N" sem palavra de status que aparece
            # ANTES de qualquer campo financeiro (Destinação, Instituição, etc.)
            # ser coletado é quase certamente uma referência dentro de uma
            # descrição (ex.: "REMANEJAMENTO DE SALDO DO ITEM 30"), não um
            # cabeçalho de item real. Nesse caso, tratamos como texto comum.
            if (
                not new_status
                and current_item is not None
                and not _item_has_financial_data(current_lines)
            ):
                current_lines.append(line)
                continue

            flush()
            current_item   = new_item
            current_status = new_status
            current_lines  = []
            continue
        if current_item is not None:
            current_lines.append(line)

    flush()
    return items


META_GERAL_LINE_RE = re.compile(r"^Meta Geral$", re.IGNORECASE)
INDICADOR_GERAL_LINE_RE = re.compile(
    r"^Indicador Geral de Resultado\b", re.IGNORECASE,
)
INDICADOR_GERAL_MARKER_RE = re.compile(
    r"Indicador Geral(?:\s+de(?:\s+Resultado)?)?\s*:?", re.IGNORECASE,
)
VALOR_REFERENCIA_RE = re.compile(r"valor de refer[eê]ncia\s*:", re.IGNORECASE)
META_ESPECIFICA_LINE_RE = re.compile(r"^META ESPEC[ÍI]FICA\s+(\d+)", re.IGNORECASE)

SECTION_LABEL_PATTERNS = [
    ("descricao_indicador", re.compile(r"^Descri[cç][aã]o do Indicador:\s*(.*)", re.IGNORECASE)),
    ("formula",             re.compile(r"^F[oó]rmula:\s*(.*)",                   re.IGNORECASE)),
    ("carteira_mjsp",       re.compile(r"^Carteira de Pol[íi]ticas do MJSP:\s*(.*)", re.IGNORECASE)),
    ("meta_pnsp",           re.compile(r"^Meta do PNSP:\s*(.*)",                 re.IGNORECASE)),
    ("meta_pesp",           re.compile(r"^Meta do PESP:\s*(.*)",                 re.IGNORECASE)),
    ("periodicidade",       re.compile(r"^Periodicidade:\s*(.*)",                 re.IGNORECASE)),
    ("fonte_ano",           re.compile(
        r"^(?:Fonte(?:/Ano)?|Valor de Refer[eê]ncia(?:/Fonte)?)\s*:\s*(.*)",
        re.IGNORECASE,
    )),
]

TECHNICAL_FIELD_FLAGS = (
    "saw_status", "saw_descricao_indicador", "saw_formula", "saw_carteira_mjsp",
)
FIELD_TO_FLAG = {
    "descricao_indicador": "saw_descricao_indicador",
    "formula":             "saw_formula",
    "carteira_mjsp":       "saw_carteira_mjsp",
}
META_PESP_CUTOFF_RE = re.compile(
    r"\b(?:Periodicidade|Fonte(?:/Ano)?|Valor de Refer[eê]ncia(?:/Fonte)?)\s*:",
    re.IGNORECASE,
)
PERIODICIDADE_FONTE_INLINE_RE = re.compile(
    r"\|\s*(?:Fonte(?:/Ano)?|Valor de Refer[eê]ncia(?:/Fonte)?)\s*:\s*(.*)$",
    re.IGNORECASE,
)


def extract_meta_geral(lines) -> str:
    for idx, line in enumerate(lines):
        if META_GERAL_LINE_RE.match(line):
            collected = []
            for next_line in lines[idx + 1:]:
                if re.match(r"^(Justificativa|Indicador Geral de Resultado|META ESPEC[ÍI]FICA)", next_line, re.IGNORECASE):
                    break
                collected.append(next_line)
            return blank_if_dash_only(" ".join(collected))
    return ""


def _extract_text_after_marker(line: str, marker_pattern) -> str:
    match = marker_pattern.search(line or "")
    if not match:
        return ""
    return blank_if_dash_only((line or "")[match.end():].lstrip(" :;-"))


def extract_indicador_geral_completo(lines) -> str:
    for idx, line in enumerate(lines):
        has_indicator_marker = bool(INDICADOR_GERAL_MARKER_RE.search(line or ""))
        is_indicator_header  = bool(INDICADOR_GERAL_LINE_RE.match(line or ""))
        is_meta_inline_indicator = bool(
            re.match(r"^Meta Geral\s*:", line or "", re.IGNORECASE)
            and has_indicator_marker
        )
        if not (is_indicator_header or is_meta_inline_indicator):
            continue
        collected = []
        inline = _extract_text_after_marker(line, INDICADOR_GERAL_MARKER_RE)
        if inline:
            collected.append(inline)
        for next_line in lines[idx + 1:]:
            if re.match(r"^META ESPEC[ÍI]FICA", next_line, re.IGNORECASE):
                break
            if re.match(r"^Meta Geral", next_line, re.IGNORECASE):
                inline_meta = _extract_text_after_marker(next_line, INDICADOR_GERAL_MARKER_RE)
                if inline_meta:
                    collected.append(inline_meta)
                    continue
                break
            if INDICADOR_GERAL_MARKER_RE.search(next_line):
                inline_next = _extract_text_after_marker(next_line, INDICADOR_GERAL_MARKER_RE)
                if inline_next:
                    collected.append(inline_next)
                continue
            if re.match(r"^(Itens da Meta|Status:)", next_line, re.IGNORECASE):
                break
            collected.append(next_line)
        indicador = blank_if_dash_only(EXAMPLE_BLOCK_RE.sub("", " ".join(collected)).strip())
        if indicador:
            return indicador
    return ""


def extract_indicador_geral_valor_referencia(lines) -> str:
    """Extrai o Valor de Referência do Indicador Geral (Meta Geral) — usado como fallback."""
    for idx, line in enumerate(lines):
        # Só busca ANTES da primeira META ESPECÍFICA
        if META_ESPECIFICA_LINE_RE.match(line):
            break
        marker_match = VALOR_REFERENCIA_RE.search(line)
        if not marker_match:
            continue
        collected = [line[marker_match.start():].strip()]
        for next_line in lines[idx + 1:]:
            if re.match(
                r"^(META ESPEC[ÍI]FICA|Descri[cç][aã]o do Indicador:|Itens da Meta|Status:)",
                next_line, re.IGNORECASE
            ):
                break
            collected.append(next_line)
        return blank_if_dash_only(" ".join(collected))
    return ""


def extract_analysis_data(lines):
    return {
        "zero_indicador_geral":   extract_indicador_geral_completo(lines),
        "one_meta_geral":         extract_meta_geral(lines),
        "three_valor_referencia": extract_indicador_geral_valor_referencia(lines),
        "sections":               extract_meta_especifica_sections(lines),
    }


def _finalize_meta_section(section):
    result = {"numero_meta": section["numero_meta"]}
    for key in (
        "meta_texto", "descricao_indicador", "formula",
        "meta_pesp", "meta_pnsp", "carteira_mjsp",
        "periodicidade", "fonte_ano",
    ):
        result[key] = blank_if_dash_only(" ".join(section.get(key, [])))
    return result


def _is_technical_meta_section(section) -> bool:
    return all(section.get(flag, False) for flag in TECHNICAL_FIELD_FLAGS)


def _dedupe_sections_keep_last(sections):
    deduped_reversed = []
    seen = set()
    for section in reversed(sections):
        meta_num = section.get("numero_meta")
        if meta_num in seen:
            continue
        seen.add(meta_num)
        deduped_reversed.append(section)
    deduped_reversed.reverse()
    return deduped_reversed


def _merge_sections_prefer_technical(all_sections, technical_sections):
    technical_by_meta = {s.get("numero_meta"): s for s in technical_sections}
    return [technical_by_meta.get(s.get("numero_meta"), s) for s in all_sections]


def _trim_meta_pesp(text: str) -> str:
    text = blank_if_dash_only(text)
    if not text:
        return ""
    cutoff = META_PESP_CUTOFF_RE.search(text)
    if cutoff:
        return blank_if_dash_only(text[:cutoff.start()])
    return blank_if_dash_only(text)


def extract_meta_especifica_sections(lines):
    sections = []
    current = None
    current_field = None

    for line in lines:
        meta_match = META_ESPECIFICA_LINE_RE.match(line)
        if meta_match:
            if current is not None:
                sections.append(current)
            current = {
                "numero_meta": int(meta_match.group(1)),
                "meta_texto": [], "descricao_indicador": [], "formula": [],
                "meta_pesp": [], "meta_pnsp": [], "carteira_mjsp": [],
                "periodicidade": [], "fonte_ano": [],
                "saw_status": False, "saw_descricao_indicador": False,
                "saw_formula": False, "saw_carteira_mjsp": False,
            }
            current_field = "meta_texto"
            continue

        if current is None:
            continue

        if re.match(r"^Status:", line, re.IGNORECASE):
            current["saw_status"] = True
            current_field = None
            continue
        if re.match(r"^Itens da Meta$", line, re.IGNORECASE):
            current_field = None
            continue
        if ITEM_RE.match(line):
            current_field = None
            continue

        matched_label = False
        for field_key, pattern in SECTION_LABEL_PATTERNS:
            match = pattern.match(line)
            if match:
                current_field = field_key
                flag_key = FIELD_TO_FLAG.get(field_key)
                if flag_key:
                    current[flag_key] = True
                content = match.group(1).strip()
                if content:
                    if field_key == "periodicidade":
                        fonte_inline = PERIODICIDADE_FONTE_INLINE_RE.search(content)
                        if fonte_inline:
                            periodicidade_value = content[:fonte_inline.start()].strip()
                            fonte_value = fonte_inline.group(1).strip()
                            if periodicidade_value:
                                current[field_key].append(periodicidade_value)
                            if fonte_value:
                                current["fonte_ano"].append(fonte_value)
                            matched_label = True
                            break
                    current[field_key].append(content)
                matched_label = True
                break
        if matched_label:
            continue

        if current_field:
            current[current_field].append(line)

    if current is not None:
        sections.append(current)

    finalized_sections = [_finalize_meta_section(s) for s in sections]
    finalized_sections = _dedupe_sections_keep_last(finalized_sections)

    technical_sections = [
        _finalize_meta_section(s) for s in sections if _is_technical_meta_section(s)
    ]
    technical_sections = _dedupe_sections_keep_last(technical_sections)

    merged_sections = _merge_sections_prefer_technical(finalized_sections, technical_sections)

    for section in merged_sections:
        section["meta_pesp"] = _trim_meta_pesp(section.get("meta_pesp", ""))
    return merged_sections


def extract_fields(item_lines):
    fields = {key: [] for key, _ in CAPTURE_PATTERNS}
    fields["acao"] = []
    fields["art"]  = []
    fields["art_num"] = ""
    current_field = None

    for line in item_lines:
        matched = False
        for stop in STOP_PATTERNS:
            if stop.match(line):
                current_field = None
                matched = True
                break
        if matched:
            continue

        action_match = ACTION_PATTERN.match(line)
        if action_match:
            current_field = "acao"
            action_body = action_match.group(1).strip()
            if action_body:
                fields[current_field].append(action_body)
            continue

        art_match = ART_PATTERN.match(line)
        if art_match:
            current_field = "art"
            art_num  = art_match.group(1)
            art_body = art_match.group(3).strip()
            if art_body:
                fields[current_field].append(art_body)
            fields["art_num"] = art_num
            continue

        for field, pattern in CAPTURE_PATTERNS:
            match = pattern.match(line)
            if match:
                current_field = field
                content = match.group(1).strip()
                if content:
                    fields[current_field].append(content)
                matched = True
                break
        if matched:
            continue

        if current_field:
            fields[current_field].append(line)

    for key in fields:
        fields[key] = blank_if_dash_only(" ".join(fields[key]))

    return fields


def _inject_reference_text(base_text: str, reference_text: str) -> str:
    if not reference_text:
        return base_text
    marker = "A referência informada foi:"
    if marker not in base_text:
        return reference_text
    return f"{base_text.split(marker, 1)[0]}{marker}\n\n\n\n{reference_text}"


def _inject_meta_text(base_text: str, marker: str, value: str) -> str:
    if not value:
        return base_text
    if marker not in base_text:
        return value
    before, after = base_text.split(marker, 1)
    suffix_idx = after.find("Existe aderência")
    suffix = f"\n\n\n\n{after[suffix_idx:].strip()}" if suffix_idx != -1 else ""
    return f"{before}{marker}\n\n\n\n{value}{suffix}"


def _inject_descricao_formula(base_text: str, descricao: str, formula: str) -> str:
    if not descricao and not formula:
        return base_text
    marker_desc    = "Descrição do Indicador:"
    marker_formula = "Fórmula:"
    if marker_desc not in base_text or marker_formula not in base_text:
        parts = []
        if descricao:
            parts.append(f"Descrição do Indicador: {descricao}")
        if formula:
            parts.append(f"Fórmula: {formula}")
        return "\n\n".join(parts)
    pre        = base_text.split(marker_desc, 1)[0]
    after_desc = base_text.split(marker_desc, 1)[1]
    after_formula = after_desc.split(marker_formula, 1)[1] if marker_formula in after_desc else ""
    suffix_idx = after_formula.find("O indicador")
    suffix     = f"\n\n{after_formula[suffix_idx:].strip()}" if suffix_idx != -1 else ""
    return f"{pre}{marker_desc}\n{descricao or ''}\n\n{marker_formula}\n{formula or ''}{suffix}"


def replace_placeholder_segment(base_text: str, token: str, value: str) -> str:
    text    = str(base_text or "")
    pattern = re.compile(re.escape(token) + r".*?" + re.escape(token), re.DOTALL)
    if not pattern.search(text):
        return text
    return pattern.sub(value or "", text)


def set_cell_font_black(ws, cell_ref: str):
    cell      = ws[cell_ref]
    font      = copy.copy(cell.font)
    font.color = "FF000000"
    cell.font  = font


def set_row_top_fonts_black(ws, row: int, start_col: int = 1, end_col: int = 12):
    for col in range(start_col, end_col + 1):
        set_cell_font_black(ws, f"{openpyxl.utils.get_column_letter(col)}{row}")


def collect_analysis_missing_cells(analysis_data):
    missing_cells = set()
    if not blank_if_dash_only(analysis_data.get("zero_indicador_geral", "")):
        missing_cells.add("F10")
    if not blank_if_dash_only(analysis_data.get("one_meta_geral", "")):
        missing_cells.add("A8")

    sections  = analysis_data.get("sections") or []
    reference = blank_if_dash_only(analysis_data.get("three_valor_referencia", ""))
    for idx, section in enumerate(sections, start=1):
        start_row = ANALYSIS_BLOCK_START_ROW + (idx - 1) * ANALYSIS_BLOCK_HEIGHT
        if not blank_if_dash_only(section.get("meta_texto", "")):
            missing_cells.add(f"A{start_row}")
        # Verifica fonte_ano da seção antes de considerar ausente
        section_fonte = blank_if_dash_only(section.get("fonte_ano", ""))
        if not section_fonte and not reference:
            missing_cells.add(f"E{start_row}")
        if not blank_if_dash_only(section.get("descricao_indicador", "")) or \
           not blank_if_dash_only(section.get("formula", "")):
            missing_cells.add(f"F{start_row}")
        if not blank_if_dash_only(section.get("meta_pesp", "")):
            missing_cells.add(f"G{start_row}")
        if not blank_if_dash_only(section.get("meta_pnsp", "")):
            missing_cells.add(f"H{start_row}")
        if not blank_if_dash_only(section.get("carteira_mjsp", "")):
            missing_cells.add(f"I{start_row}")
    return sorted(missing_cells)


def build_material(bem, descricao, destinacao):
    parts = []
    if bem:
        parts.append(f"Bem/Serviço: {bem}")
    if descricao:
        parts.append(f"Descrição: {descricao}")
    if destinacao:
        parts.append(f"Destinação: {destinacao}")
    return " | ".join(parts)


def _count_analysis_blocks(ws) -> int:
    block_height     = _infer_analysis_block_height(ws)
    items_title_row  = _find_items_title_row(ws)
    if items_title_row and items_title_row > ANALYSIS_BLOCK_START_ROW and block_height > 0:
        return max(1, (items_title_row - ANALYSIS_BLOCK_START_ROW) // block_height)
    return 1


def _find_items_title_row(ws):
    for row in range(1, ws.max_row + 1):
        value = normalize(str(ws.cell(row=row, column=1).value or "")).upper()
        if value == "ITENS DE CONTRATAÇÃO":
            return row
    return None


def _infer_analysis_block_height(ws) -> int:
    items_title_row = _find_items_title_row(ws)
    for merged in ws.merged_cells.ranges:
        if (
            merged.min_col == ANALYSIS_BLOCK_START_COL
            and merged.max_col == ANALYSIS_BLOCK_START_COL
            and merged.min_row == ANALYSIS_BLOCK_START_ROW
        ):
            h = merged.max_row - merged.min_row + 1
            if h > 0:
                return h
    if items_title_row and items_title_row > ANALYSIS_BLOCK_START_ROW:
        compact_height = items_title_row - ANALYSIS_BLOCK_START_ROW
        if 1 <= compact_height <= ANALYSIS_BLOCK_HEIGHT:
            return compact_height
    return ANALYSIS_BLOCK_HEIGHT


def _copy_analysis_block(ws, src_start_row: int, dst_start_row: int, block_height: int):
    for row_offset in range(block_height):
        src_row = src_start_row + row_offset
        dst_row = dst_start_row + row_offset
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
        for col in range(ANALYSIS_BLOCK_START_COL, ANALYSIS_BLOCK_END_COL + 1):
            src_cell = ws.cell(src_row, col)
            dst_cell = ws.cell(dst_row, col)
            dst_cell.value        = src_cell.value
            dst_cell.font         = copy.copy(src_cell.font)
            dst_cell.fill         = copy.copy(src_cell.fill)
            dst_cell.border       = copy.copy(src_cell.border)
            dst_cell.alignment    = copy.copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection   = copy.copy(src_cell.protection)

    shift = dst_start_row - src_start_row
    template_merges = [
        rng for rng in list(ws.merged_cells.ranges)
        if (
            rng.min_row >= src_start_row
            and rng.max_row < src_start_row + block_height
            and rng.min_col >= ANALYSIS_BLOCK_START_COL
            and rng.max_col <= ANALYSIS_BLOCK_END_COL
        )
    ]
    for rng in template_merges:
        ws.merge_cells(
            start_row=rng.min_row + shift, start_column=rng.min_col,
            end_row=rng.max_row + shift,   end_column=rng.max_col,
        )


def _unmerge_analysis_block_region(ws, block_start_row: int, block_height: int):
    block_end_row = block_start_row + block_height - 1
    to_unmerge = [
        str(rng) for rng in list(ws.merged_cells.ranges)
        if (
            rng.min_row >= block_start_row and rng.max_row <= block_end_row
            and rng.min_col >= ANALYSIS_BLOCK_START_COL
            and rng.max_col <= ANALYSIS_BLOCK_END_COL
        )
    ]
    for rng in to_unmerge:
        ws.unmerge_cells(rng)


def _shift_row_dimensions_on_insert(ws, insert_at: int, amount: int):
    if amount <= 0:
        return
    rows_to_shift = sorted(
        [r for r in ws.row_dimensions if isinstance(r, int) and r >= insert_at],
        reverse=True,
    )
    for row_idx in rows_to_shift:
        dim_copy       = copy.copy(ws.row_dimensions[row_idx])
        dim_copy.index = row_idx + amount
        ws.row_dimensions[row_idx + amount] = dim_copy
        del ws.row_dimensions[row_idx]


def _ranges_overlap(a, b) -> bool:
    return not (a[2] < b[0] or b[2] < a[0] or a[3] < b[1] or b[3] < a[1])


def _insert_rows_preserving_merges(ws, insert_at: int, amount: int):
    if amount <= 0:
        return
    original_ranges = [
        (rng.min_row, rng.min_col, rng.max_row, rng.max_col)
        for rng in list(ws.merged_cells.ranges)
    ]
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))
    ws.insert_rows(insert_at, amount)
    _shift_row_dimensions_on_insert(ws, insert_at, amount)
    rebuilt = []

    def add_range(min_row, min_col, max_row, max_col):
        if min_row > max_row or min_col > max_col:
            return
        if min_row == max_row and min_col == max_col:
            return
        candidate = (min_row, min_col, max_row, max_col)
        for existing in rebuilt:
            if _ranges_overlap(candidate, existing):
                return
        rebuilt.append(candidate)

    for min_row, min_col, max_row, max_col in original_ranges:
        if max_row < insert_at:
            add_range(min_row, min_col, max_row, max_col)
        elif min_row >= insert_at:
            add_range(min_row + amount, min_col, max_row + amount, max_col)
        else:
            add_range(min_row, min_col, insert_at - 1, max_col)
            add_range(insert_at + amount, min_col, max_row + amount, max_col)

    for min_row, min_col, max_row, max_col in rebuilt:
        ws.merge_cells(
            start_row=min_row, start_column=min_col,
            end_row=max_row,   end_column=max_col,
        )


def _ensure_analysis_blocks(ws, required_blocks: int):
    block_height    = _infer_analysis_block_height(ws)
    existing_blocks = _count_analysis_blocks(ws)
    if required_blocks <= existing_blocks:
        return
    items_title_row       = _find_items_title_row(ws)
    extra_blocks          = required_blocks - existing_blocks
    additional_rows_needed = extra_blocks * block_height
    insert_at             = ANALYSIS_BLOCK_START_ROW + existing_blocks * block_height
    minimum_gap_rows      = 1

    if items_title_row and items_title_row <= insert_at:
        rows_to_shift_items = (insert_at - items_title_row) + minimum_gap_rows
        _insert_rows_preserving_merges(ws, items_title_row, rows_to_shift_items)
        items_title_row += rows_to_shift_items

    reusable_gap_rows = 0
    if items_title_row and items_title_row > insert_at:
        reusable_gap_rows = max(0, (items_title_row - insert_at) - minimum_gap_rows)
    rows_to_insert = max(0, additional_rows_needed - reusable_gap_rows)
    _insert_rows_preserving_merges(ws, insert_at, rows_to_insert)

    for block_idx in range(existing_blocks + 1, required_blocks + 1):
        dst_start_row = ANALYSIS_BLOCK_START_ROW + (block_idx - 1) * block_height
        _unmerge_analysis_block_region(ws, dst_start_row, block_height)
        _copy_analysis_block(ws, ANALYSIS_BLOCK_START_ROW, dst_start_row, block_height)


def fill_analysis_template(ws, lines):
    analysis_data    = extract_analysis_data(lines)
    indicador_geral  = analysis_data["zero_indicador_geral"]
    meta_geral       = analysis_data["one_meta_geral"]
    valor_referencia = analysis_data["three_valor_referencia"]  # fallback global
    sections         = analysis_data["sections"]

    base_a8     = str(ws["A8"].value or "")
    a8_replaced = replace_placeholder_segment(base_a8, "1*", meta_geral)
    ws["A8"]    = a8_replaced if a8_replaced != base_a8 else (meta_geral or base_a8)
    set_cell_font_black(ws, "A8")

    base_f10     = str(ws["F10"].value or "")
    f10_replaced = replace_placeholder_segment(base_f10, "0*", indicador_geral)
    ws["F10"]    = f10_replaced if f10_replaced != base_f10 else (indicador_geral or base_f10)
    set_cell_font_black(ws, "F10")

    if not sections:
        return

    block_height = _infer_analysis_block_height(ws)
    _ensure_analysis_blocks(ws, len(sections))

    for idx in range(2, len(sections) + 1):
        start_row = ANALYSIS_BLOCK_START_ROW + (idx - 1) * block_height
        _unmerge_analysis_block_region(ws, start_row, block_height)
        _copy_analysis_block(ws, ANALYSIS_BLOCK_START_ROW, start_row, block_height)

    for idx, section in enumerate(sections, start=1):
        start_row     = ANALYSIS_BLOCK_START_ROW + (idx - 1) * block_height
        meta_text_raw = section.get("meta_texto", "")
        meta_text     = re.sub(r"^\d+\s*-\s*", "", meta_text_raw).strip()
        two_meta_texto = f"{idx} - {meta_text}" if meta_text else ""

        # Coluna A — texto da meta específica
        cell_a    = f"A{start_row}"
        base_a    = str(ws[cell_a].value or "")
        a_replaced = replace_placeholder_segment(base_a, "2*", two_meta_texto)
        ws[cell_a] = a_replaced if a_replaced != base_a else two_meta_texto

        # ── CORREÇÃO PRINCIPAL ──────────────────────────────────────────────
        # Coluna E — Valor de Referência/Fonte da própria meta específica.
        # Usa o campo "fonte_ano" extraído da seção; só recorre ao valor da
        # Meta Geral se a seção não tiver seu próprio valor de referência.
        cell_e         = f"E{start_row}"
        base_e         = str(ws[cell_e].value or "")
        section_fonte  = blank_if_dash_only(section.get("fonte_ano", ""))
        valor_ref_section = section_fonte if section_fonte else valor_referencia
        e_replaced     = replace_placeholder_segment(base_e, "3*", valor_ref_section)
        if e_replaced == base_e:
            e_replaced = _inject_reference_text(base_e, valor_ref_section)
        ws[cell_e] = e_replaced
        # ───────────────────────────────────────────────────────────────────

        # Coluna F — Descrição do Indicador + Fórmula
        cell_f    = f"F{start_row}"
        base_f    = str(ws[cell_f].value or "")
        f_replaced = replace_placeholder_segment(base_f, "4*", section.get("descricao_indicador", ""))
        f_replaced = replace_placeholder_segment(f_replaced, "5*", section.get("formula", ""))
        if f_replaced == base_f:
            f_replaced = _inject_descricao_formula(
                base_f, section.get("descricao_indicador", ""), section.get("formula", ""),
            )
        ws[cell_f] = f_replaced

        # Coluna G — Meta do PESP
        cell_g    = f"G{start_row}"
        base_g    = str(ws[cell_g].value or "")
        g_replaced = replace_placeholder_segment(base_g, "6*", section.get("meta_pesp", ""))
        if g_replaced == base_g:
            g_replaced = _inject_meta_text(base_g, "A Meta informada foi:", section.get("meta_pesp", ""))
        ws[cell_g] = g_replaced

        # Coluna H — Meta do PNSP
        cell_h    = f"H{start_row}"
        base_h    = str(ws[cell_h].value or "")
        h_replaced = replace_placeholder_segment(base_h, "7*", section.get("meta_pnsp", ""))
        if h_replaced == base_h:
            h_replaced = _inject_meta_text(base_h, "A Meta informada foi:", section.get("meta_pnsp", ""))
        ws[cell_h] = h_replaced

        # Coluna I — Carteira de Políticas do MJSP
        cell_i    = f"I{start_row}"
        base_i    = str(ws[cell_i].value or "")
        i_replaced = replace_placeholder_segment(base_i, "8*", section.get("carteira_mjsp", ""))
        if i_replaced == base_i:
            i_replaced = _inject_meta_text(base_i, "A política informada foi:", section.get("carteira_mjsp", ""))
        ws[cell_i] = i_replaced

        set_row_top_fonts_black(ws, start_row, 1, 12)


def fill_worksheet(ws, rows, header_map, start_row=3):
    max_col = max(header_map.values()) if header_map else ws.max_column
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, max_col=max_col):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None

    style_template_row = start_row if start_row <= ws.max_row else None
    style_template_has_custom_style = False
    if style_template_row is not None and header_map:
        style_template_has_custom_style = any(
            ws.cell(style_template_row, col_idx).style_id != 0
            for col_idx in header_map.values()
        )

    for idx, row_data in enumerate(rows, start=start_row):
        if (
            style_template_row is not None
            and style_template_has_custom_style
            and header_map
            and idx != style_template_row
            and all(ws.cell(idx, col_idx).style_id == 0 for col_idx in header_map.values())
        ):
            for col_idx in header_map.values():
                ws.cell(idx, col_idx)._style = copy.copy(
                    ws.cell(style_template_row, col_idx)._style
                )
            template_height = ws.row_dimensions[style_template_row].height
            if template_height is not None:
                ws.row_dimensions[idx].height = template_height
        for header, col_idx in header_map.items():
            cell = ws.cell(row=idx, column=col_idx)
            if isinstance(cell, MergedCell):
                continue
            cell.value = row_data.get(header, "")


def get_template_header_info(template_path: Path):
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    headers, header_map = get_header_info_from_ws(ws, 2)
    if not header_map:
        headers = OUTPUT_HEADERS[:]
        header_map = {}
        for idx, header in enumerate(headers):
            key = _header_key(header)
            if key not in header_map:
                header_map[key] = idx + 1
    return headers, header_map


def find_items_table_header_row(ws):
    for row in range(1, ws.max_row + 1):
        value = normalize(str(ws.cell(row=row, column=1).value or "")).upper()
        if value == "ITENS DE CONTRATAÇÃO":
            return row + 1
    return None


def get_analysis_items_header_info(template_path: Path):
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    header_row = find_items_table_header_row(ws)
    if not header_row:
        return None, [], {}
    headers, header_map = get_header_info_from_ws(ws, header_row)
    return header_row, headers, header_map


def update_action_header(ws, rows, header_map, art_num_preferred=None, header_row=2):
    col_idx = header_map.get(ACTION_HEADER_KEY)
    if not col_idx or not rows:
        return
    art_num = art_num_preferred or rows[0].get(ACTION_HEADER_NUM_KEY)
    if not art_num or str(art_num) not in {"6", "7", "8"}:
        return
    ws.cell(row=header_row, column=col_idx,
            value=f"Ação conforme Art. {art_num}º da portaria nº 685")


def generate_excel_bytes(
    template_path: Path, rows, header_map,
    art_num_preferred=None, source_lines=None,
) -> bytes:
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    if is_analysis_template_sheet(ws):
        fill_analysis_template(ws, source_lines or [])
        header_row = find_items_table_header_row(ws)
        if header_row and rows:
            _, items_header_map = get_header_info_from_ws(ws, header_row)
            update_action_header(ws, rows, items_header_map,
                                 art_num_preferred=art_num_preferred, header_row=header_row)
            fill_worksheet(ws, rows, items_header_map, start_row=header_row + 1)
    else:
        update_action_header(ws, rows, header_map, art_num_preferred=art_num_preferred)
        fill_worksheet(ws, rows, header_map)
    ws.sheet_view.topLeftCell              = "A1"
    ws.sheet_view.selection[0].activeCell  = "A1"
    ws.sheet_view.selection[0].sqref       = "A1"
    ws.sheet_view.zoomScale                = 100
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_rows(parsed_items, header_map):
    has_descricao         = "Descrição"       in header_map
    has_destinacao        = "Destinação"      in header_map
    has_quantidade_unidade = "Quantidade/Unidade" in header_map
    has_valor_status      = "Valor/Status"    in header_map
    has_unidade_col       = "Unidade de Medida" in header_map
    has_status_col        = "Status do Item"  in header_map
    rows = []
    for item in parsed_items:
        fields = extract_fields(item["lines"])
        material = (
            fields["bem"]
            if has_descricao or has_destinacao
            else build_material(fields["bem"], fields["descricao"], fields["destinacao"])
        )
        valor_total  = format_currency(fields["valor_total"])
        quantidade   = parse_int(fields["quantidade"])
        unidade      = fields["unidade"]
        status_item  = item.get("status") or "Planejado"

        quantidade_unidade = ""
        if has_quantidade_unidade and not has_unidade_col:
            if quantidade != "" and unidade:
                quantidade_unidade = f"{quantidade} {unidade}"
            elif quantidade != "":
                quantidade_unidade = str(quantidade)
            elif unidade:
                quantidade_unidade = unidade

        valor_status = ""
        if has_valor_status and not has_status_col:
            if valor_total and status_item:
                valor_status = f"{valor_total} | {status_item}"
            elif valor_total:
                valor_status = valor_total
            elif status_item:
                valor_status = status_item

        rows.append({
            "Número da Meta Específica":                 item["meta"],
            "Número do Item":                            item["item"],
            ACTION_HEADER_KEY:                           fields["acao"] or fields["art"],
            ACTION_HEADER_NUM_KEY:                       fields["art_num"],
            "Material/Serviço":                          material,
            "Descrição":                                 fields["descricao"] if has_descricao else "",
            "Destinação":                                fields["destinacao"] if has_destinacao else "",
            "Instituição":                               fields["instituicao"],
            "Natureza da Despesa":                       fields["natureza"],
            "Quantidade Planejada":                      quantidade,
            "Unidade de Medida":                         fields["unidade"],
            "Quantidade/Unidade":                        quantidade_unidade,
            "Valor Planejado Total":                     valor_total,
            "Status do Item":                            status_item,
            "Valor/Status":                              valor_status,
        })
    return rows
